import zipfile
from script_os import ZIP_DIR
from pypdf import PdfReader
from openpyxl import load_workbook
import csv

def test_check_pdf_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("test_file_pdf.pdf") as pdf_file:
            pd = PdfReader(pdf_file)
            num_pages = len(pd.pages)

            text = pd.pages[0].extract_text()

            assert num_pages == 484, "Количество страниц не сходится"
            assert text == 'pytest Documentation\n''Release 0.1\n''holger krekel, trainer and consultant, https://merlinux.eu/\n''Jun 04, 2024'

def test_check_pdf_file_xlsx():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("test_file_xlsx.xlsx") as xlsx_file:
            workbook = load_workbook(filename=xlsx_file)
            sheet = workbook.active
            workbook_value = (sheet.cell(row=3, column=2).value)

            assert workbook_value == 'OU001'

def test_check_pdf_file_csv():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("test_file_csv.csv") as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]

            assert second_row[0][2] == 'i'
            assert int(second_row[0][0]) == 1
